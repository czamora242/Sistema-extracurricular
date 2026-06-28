
from pathlib import Path
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QColor,QFont
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
    QTableWidgetItem, QPushButton, QLineEdit, QMessageBox,
    QHeaderView, QFrame, QLabel, QFileDialog, QComboBox,
    QDialog,QMenu
)

from services.docente_service import DocenteService
from services.usuario_service import UsuarioService
from services.auth_service import SesionUsuario
from ui.docentes.dialogo_crear_docente import DialogoCrearDocente
from ui.docentes.dialogo_editar_docente import DialogoEditarDocente


class PanelGestionarDocentes(QWidget):
    """
    Panel completo para gestionar docentes del sistema.
    CORREGIDO: sin duplicaciones, mejor manejo de IDs.
    """

    def __init__(self, sesion_usuario: SesionUsuario, parent=None):
        super().__init__(parent)
        self.sesion_usuario = sesion_usuario
        self.docente_seleccionado_id = None  # Guardar solo el ID
        self.docentes_datos = []  # Caché de docentes
        self._timer   = QTimer()    
        self._timer.setSingleShot(True)
        self._timer.setInterval(350)   # ms de espera antes de buscar
        self._timer.timeout.connect(self._filtrar_tabla)
        self._construir_ui()
        self._cargar_docentes()

    
    def _construir_ui(self) -> None:
        raiz = QVBoxLayout(self)
        raiz.setContentsMargins(24,20,24,16)
        raiz.setSpacing(0)

        raiz.addLayout(self._cabecera())
        raiz.addSpacing(14)

        raiz.addWidget(self._barra_busqueda())
        raiz.addSpacing(12)

        self.tabla = self._construir_tabla()
        raiz.addWidget(self.tabla, 1)

        raiz.addSpacing(10)
        raiz.addLayout(self._pie())

    def _cabecera(self):
        lay = QHBoxLayout()

        lbl = QLabel("Docentes")
        lbl.setObjectName("lbl_titulo_modulo")
        font = QFont()
        font.setPointSize(20)
        font.setWeight(QFont.Weight.Bold)
        lbl.setFont(font)
        lay.addWidget(lbl)
        lay.addStretch()


        self.btn_importar = QPushButton("↑ Importar Excel")
        self.btn_importar.setFixedHeight(36)
        self.btn_importar.setToolTip(
            "Importar docentes masivamente desde un archivo .xlsx")
        self.btn_importar.clicked.connect(self._importar_excel)
        lay.addWidget(self.btn_importar)

        lay.addSpacing(8)

        self.btn_nuevo = QPushButton("+ Nuevo docente")
        self.btn_nuevo.setFixedHeight(36)
        self.btn_nuevo.clicked.connect(self._abrir_crear_docente)

        # Solo admin/operador pueden crear
        if self.sesion_usuario.rol_nombre == "Docente":
            self.btn_nuevo.setEnabled(False)
            self.btn_nuevo.setToolTip("Sin permiso para crear docentes")
        lay.addWidget(self.btn_nuevo)

        return lay   
    
    def _barra_busqueda(self):
        frame = QFrame()
        lay = QHBoxLayout(frame)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(8)

        self.inp_buscar = QLineEdit()
        self.inp_buscar.setPlaceholderText("🔍 Buscar por nombre, DNI o especialidad...")
        self.inp_buscar.textChanged.connect(self._timer.start)
        lay.addWidget(self.inp_buscar)

        self.btn_limpiar = QPushButton("✕ Limpiar")
        self.btn_limpiar.clicked.connect(lambda: self.inp_buscar.clear())
        lay.addWidget(self.btn_limpiar)

        return frame
    
    def _construir_tabla(self):
        tabla = QTableWidget()
        tabla.setColumnCount(8)
        tabla.setHorizontalHeaderLabels([
            "#",
            "DNI",
            "Nombre",
            "Apellidos",
            "Especialidad",
            "Email",
            "Usuario",
            "Estado"
        ])

        tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        tabla.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        tabla.setAlternatingRowColors(True)
        tabla.verticalHeader().setVisible(False)
        tabla.setShowGrid(False)
        tabla.setSortingEnabled(True)
        tabla.itemSelectionChanged.connect(self._on_seleccionar_docente)
        tabla.doubleClicked.connect(self._abrir_editar_docente) 
        tabla.verticalHeader().setDefaultSectionSize(38)

        #Cabecera de la tabla
        header = tabla.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.Stretch)
        header.setSectionResizeMode(5, QHeaderView.Stretch)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)
        return tabla
    
    def _pie(self):
        lay = QHBoxLayout()
        self.lbl_conteo = QLabel()
        lay.addWidget(self.lbl_conteo)
        lay.addStretch()

        self.btn_editar = QPushButton("✎ Editar")
        self.btn_editar.clicked.connect(self._abrir_editar_docente)
        self.btn_editar.setEnabled(False)
        lay.addWidget(self.btn_editar)

        self.btn_vincular = QPushButton("🔗 Vincular")
        self.btn_vincular.clicked.connect(self._vincular_usuario)
        self.btn_vincular.setEnabled(False)
        lay.addWidget(self.btn_vincular)

        self.btn_desvincular = QPushButton("🔗 Desvincular")
        self.btn_desvincular.clicked.connect(self._desvincular_usuario)
        self.btn_desvincular.setEnabled(False)
        lay.addWidget(self.btn_desvincular)

        self.btn_cambiar_estado = QPushButton("Cambiar estado")
        self.btn_cambiar_estado.clicked.connect(self._cambiar_estado)
        self.btn_cambiar_estado.setEnabled(False)
        lay.addWidget(self.btn_cambiar_estado)

        self.btn_eliminar = QPushButton("🗑 Eliminar")
        self.btn_eliminar.clicked.connect(self._eliminar_docente)
        self.btn_eliminar.setEnabled(False)
        lay.addWidget(self.btn_eliminar)

        return lay
    # ══════════════════════════════════════════════════════════════
    # CARGAR Y FILTRAR DATOS
    # ══════════════════════════════════════════════════════════════

    def _cargar_docentes(self) -> None:
        """Carga la lista de docentes de la BD."""
        resultado = DocenteService.listar(activos_solo=False)

        if not resultado.ok:
            QMessageBox.warning(self, "Error", resultado.mensaje)
            return

        # Guardar datos en caché
        self.docentes_datos = resultado.lista or []
        
        # Limpiar selección
        self.docente_seleccionado_id = None
        self.tabla.clearSelection()
        
        # Actualizar tabla
        self._actualizar_tabla()

    def _actualizar_tabla(self) -> None:
        """Actualiza la tabla SIN DUPLICACIONES."""
        # Importante: no mezclar datos antiguos con nuevos
        self.tabla.setSortingEnabled(False)
        self.tabla.setRowCount(len(self.docentes_datos))

        for row, docente in enumerate(self.docentes_datos):
            # Crear items con datos correctos
            items = [
                QTableWidgetItem(str(docente["id"])),
                QTableWidgetItem(docente["dni"]),
                QTableWidgetItem(docente["nombre"]),
                QTableWidgetItem(docente["apellidos"]),
                QTableWidgetItem(docente["especialidad"] or "—"),
                QTableWidgetItem(docente["email"] or "—"),
                QTableWidgetItem(docente["usuario"] if docente["usuario"] != "Sin usuario" else "—"),
                QTableWidgetItem("Activo" if docente["activo"] else "Inactivo")
            ]

            # Configurar items
            for col, item in enumerate(items):
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if col == 0:
                    item.setData(Qt.ItemDataRole.UserRole, docente["id"])

                
                # Colorear estado
                if col == 7:
                    if item.text() == "Activo":
                        item.setForeground(QColor("#1D9E75"))
                        item.setBackground(QColor("#E7F7F1"))
                    else:
                        item.setForeground(QColor("#8c7b6e"))
                        item.setBackground(QColor("#F4ECE7"))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                
                if col == 0:
                    item.setData(Qt.ItemDataRole.UserRole, docente["id"])

                # Establecer item en tabla
                self.tabla.setItem(row, col, item)
                
        self.tabla.setSortingEnabled(True)

        # Actualizar contador
        n = len(self.docentes_datos)
        sufijo = "" if n == 1 else "s"
        self.lbl_conteo.setText(f"Mostrando {n} docente{sufijo}")

    def _filtrar_tabla(self) -> None:
        """Filtra la tabla según búsqueda."""
        buscar = self.inp_buscar.text().lower()

        for row in range(len(self.docentes_datos)):
            docente = self.docentes_datos[row]
            coincide = (
                buscar in docente["dni"].lower() or
                buscar in docente["nombre"].lower() or
                buscar in (docente["especialidad"] or "").lower()
            )
            self.tabla.setRowHidden(row, not coincide)

    # ══════════════════════════════════════════════════════════════
    # SELECCIÓN Y ACCIONES
    # ══════════════════════════════════════════════════════════════

    def _obtener_docente_actual(self) -> dict:
        """
        Obtiene el docente actualmente seleccionado.
        IMPORTANTE: busca por ID para evitar duplicaciones.
        """
        if self.docente_seleccionado_id is None:
            return None
        
        return next(
            (d for d in self.docentes_datos if d["id"] == self.docente_seleccionado_id),
            None
        )
    def _docente_id_seleccionado(self):
        filas = self.tabla.selectedItems()
        if not filas:
            return None
        fila = self.tabla.currentRow()
        item = self.tabla.item(fila,0)
        return item.data(Qt.ItemDataRole.UserRole) if item else None

    def _on_seleccionar_docente(self) -> None:
        """Se ejecuta cuando selecciona un docente en la tabla."""
        rows = self.tabla.selectedIndexes()
        if not rows:
            self.docente_seleccionado_id = None
            self._desabilitar_botones()
            return

        # Obtener el ID de la fila seleccionada
        row = rows[0].row()
        self.docente_seleccionado_id = self._docente_id_seleccionado()
        docente = self._obtener_docente_actual()

        if not docente:
            self.docente_seleccionado_id = None
            self._desabilitar_botones()
            return

        # Habilitar botones apropiados
        self.btn_editar.setEnabled(True)
        self.btn_vincular.setEnabled(True)
        self.btn_cambiar_estado.setEnabled(True)
        self.btn_eliminar.setEnabled(True)

        # Desbloquear vincular solo si no tiene usuario
        self.btn_vincular.setEnabled(docente["usuario"] == "Sin usuario")
        self.btn_desvincular.setEnabled(docente["usuario"] != "Sin usuario")

    def _desabilitar_botones(self) -> None:
        """Desabilita todos los botones de acción."""
        self.btn_editar.setEnabled(False)
        self.btn_vincular.setEnabled(False)
        self.btn_cambiar_estado.setEnabled(False)
        self.btn_eliminar.setEnabled(False)

    def _abrir_crear_docente(self) -> None:
        """Abre el dialog para crear docente."""
        dialogo = DialogoCrearDocente(
            sesion_usuario=self.sesion_usuario,
            parent=self
        )

        if dialogo.exec():
            self._cargar_docentes()
            QMessageBox.information(self, "Éxito", "✓ Docente creado correctamente.")

    def _abrir_editar_docente(self) -> None:
        """Abre el dialog para editar docente."""
        if self.docente_seleccionado_id is None:
            QMessageBox.warning(self, "Error", "Selecciona un docente para editar.")
            return

        dialogo = DialogoEditarDocente(
            docente_id=self.docente_seleccionado_id,
            sesion_usuario=self.sesion_usuario,
            parent=self
        )

        if dialogo.exec():
            self._cargar_docentes()
            QMessageBox.information(self, "Éxito", "✓ Docente actualizado correctamente.")

    def _vincular_usuario(self) -> None:
        """Abre el nuevo dialog de vinculación."""
        from ui.docentes.dialogo_vincular_usuario_docente import DialogoVincularDocenteUsuario
        
        dialogo = DialogoVincularDocenteUsuario(
            sesion_usuario=self.sesion_usuario,
            parent=self
        )
        dialogo.exec()
        self._cargar_docentes()  

    def _desvincular_usuario(self) -> None:
        """Desvincula el docente del usuario."""
        docente = self._obtener_docente_actual()
        if not docente:
            return

        resultado = QMessageBox.question(
            self,
            "Confirmar",
            f"¿Desvincular docente '{docente['nombre']}' de su usuario?"
        )

        if resultado != QMessageBox.Yes:
            return

        res = DocenteService.desvincular_usuario(
            docente_id=docente["id"],
            usuario_editor_id=self.sesion_usuario.usuario_id
        )

        if res.ok:
            QMessageBox.information(self, "Éxito", res.mensaje)
            self._cargar_docentes()
        else:
            QMessageBox.warning(self, "Error", res.mensaje)

        self._cargar_docentes()

    def _cambiar_estado(self) -> None:
        """Muestra un menú para activar o desactivar el docente."""

        docente = self._obtener_docente_actual()
        if not docente:
            return

        resultado = DocenteService.obtener_por_id(docente["id"])
        if not resultado.ok:
            QMessageBox.warning(self, "Error", resultado.mensaje)
            return
        datos = resultado.datos
        menu = QMenu(self)

        if datos["activo"]:
            accion = menu.addAction("Desactivar docente")
            accion.triggered.connect(
                lambda: self._confirmar_cambio_estado(docente["id"], False)
            )
        else:
            accion = menu.addAction("Activar docente")
            accion.triggered.connect(
                lambda: self._confirmar_cambio_estado(docente["id"], True)
            )

        menu.exec(
            self.btn_cambiar_estado.mapToGlobal(
                self.btn_cambiar_estado.rect().bottomLeft()
            )
        )

    def _confirmar_cambio_estado(self,docente_id: int,activar: bool) -> None:
        """Confirma el cambio de estado."""

        resultado = DocenteService.obtener_por_id(docente_id)

        if not resultado.ok:
            QMessageBox.warning(self, "Error", resultado.mensaje)
            return

        datos = resultado.datos
        accion = "activar" if activar else "desactivar"
        resp = QMessageBox.question(
            self,
            "Confirmar",
            f"¿Desea {accion} al docente\n\n"
            f"{datos['nombre_completo']}?"
        )

        if resp != QMessageBox.Yes:
            return

        if activar:
            res = DocenteService.activar(
                docente_id=docente_id,
                usuario_editor_id=self.sesion_usuario.usuario_id
            )
        else:
            res = DocenteService.desactivar(
                docente_id=docente_id,
                usuario_editor_id=self.sesion_usuario.usuario_id
            )

        if res.ok:
            QMessageBox.information(self, "Éxito", res.mensaje)
            self._cargar_docentes()
        else:
            QMessageBox.warning(self, "Error", res.mensaje)


    def _eliminar_docente(self) -> None:
        """Elimina el docente seleccionado."""
        docente = self._obtener_docente_actual()
        if not docente:
            return

        resultado = QMessageBox.warning(
            self,
            "⚠️ Advertencia",
            f"¿Estás seguro de eliminar a '{docente['nombre']}'?\n\nEsta acción NO se puede deshacer.",
            QMessageBox.Yes | QMessageBox.No
        )

        if resultado != QMessageBox.Yes:
            return

        res = DocenteService.eliminar(
            docente_id=docente["id"],
            usuario_editor_id=self.sesion_usuario.usuario_id
        )

        if res.ok:
            QMessageBox.information(self, "Éxito", res.mensaje)
            self._cargar_docentes()
        else:
            QMessageBox.warning(self, "Error", res.mensaje)

    # ══════════════════════════════════════════════════════════════
    # IMPORTAR DESDE EXCEL
    # ══════════════════════════════════════════════════════════════

    def _importar_excel(self) -> None:
        """Importa docentes desde un archivo Excel."""
        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Selecciona archivo de docentes",
            "",
            "Archivos Excel (*.xlsx *.xls);;Todos los archivos (*)"
        )

        if not archivo:
            return

        try:
            import openpyxl
            from openpyxl import load_workbook
        except ImportError:
            QMessageBox.warning(
                self,
                "Error",
                "Se requiere openpyxl.\nInstala con: pip install openpyxl"
            )
            return

        try:
            wb = load_workbook(archivo)
            ws = wb.active

            docentes_creados = 0
            docentes_error = 0
            errores = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue

                try:
                    dni = str(row[0]).strip()
                    nombres = str(row[1]).strip() if row[1] else ""
                    apellidos = str(row[2]).strip() if row[2] else ""
                    especialidad = str(row[3]).strip() if row[3] else None
                    email = str(row[4]).strip() if row[4] else None
                    telefono = str(row[5]).strip() if row[5] else None

                    if not nombres or not apellidos:
                        errores.append(f"Fila {row_idx}: Nombres y apellidos obligatorios")
                        docentes_error += 1
                        continue

                    resultado = DocenteService.crear(
                        dni=dni,
                        nombres=nombres,
                        apellidos=apellidos,
                        especialidad=especialidad,
                        email_institucional=email,
                        telefono=telefono,
                        usuario_creador_id=self.sesion_usuario.usuario_id
                    )

                    if resultado.ok:
                        docentes_creados += 1
                    else:
                        errores.append(f"Fila {row_idx}: {resultado.mensaje}")
                        docentes_error += 1

                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")
                    docentes_error += 1

            mensaje = f"✓ Docentes creados: {docentes_creados}\n✗ Errores: {docentes_error}"
            if errores and len(errores) <= 5:
                mensaje += "\n\nErrores:\n" + "\n".join(errores[:5])
            elif errores:
                mensaje += f"\n\n(Mostrando primeros 5 de {len(errores)} errores)"

            QMessageBox.information(self, "Importación completada", mensaje)
            self._cargar_docentes()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al leer archivo:\n{str(e)}")