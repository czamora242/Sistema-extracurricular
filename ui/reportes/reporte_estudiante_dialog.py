"""
ui/reportes/reporte_estudiante_mejorado_dialog.py   ──   Reporte Estudiante Mejorado
═════════════════════════════════════════════════════════════════════════════════════

Incluye:
  ✅ Filtros por ciclo, nombre, carrera, taller
  ✅ Búsqueda en vivo
  ✅ Exportación a Excel y PDF
  ✅ Contabilización correcta (P, J, A)
"""

from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QFont, QColor
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTableWidget, QTableWidgetItem, QMessageBox, QFileDialog,
    QComboBox, QLineEdit, QHeaderView, QFrame, QScrollArea,
    QWidget, QProgressBar
)

from services.reporte_service import ReporteService
from utils.exportar_excel import ExportarExcel
from utils.exportar_pdf import ExportarPDF


class ReporteEstudianteDialog(QDialog):
    """Diálogo mejorado de reporte de estudiante con filtros."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.reporte_service = ReporteService()
        self.estudiantes_filtrados = []
        self.reporte_actual = None
        self._timer_busqueda = QTimer(singleShot=True, interval=300)
        self._timer_busqueda.timeout.connect(self._ejecutar_busqueda)
        
        self._configurar_ventana()
        self._construir_ui()
        self._cargar_datos_iniciales()
    
    def _configurar_ventana(self) -> None:
        """Configura la ventana del dialog."""
        self.setWindowTitle("Reporte de Estudiante")
        self.setMinimumSize(1200, 750)
        self.setModal(True)
    
    def _construir_ui(self) -> None:
        """Construye la interfaz."""
        layout = QVBoxLayout()
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)
        
        # ── TÍTULO ─────────────────────────────────────────────────
        lbl_titulo = QLabel("👤 Reporte de Estudiante")
        font_titulo = QFont()
        font_titulo.setPointSize(16)
        font_titulo.setBold(True)
        lbl_titulo.setFont(font_titulo)
        layout.addWidget(lbl_titulo)
        
        # ── FILTROS ────────────────────────────────────────────────
        layout.addWidget(self._crear_panel_filtros())
        
        # ── TABLA DE ESTUDIANTES ───────────────────────────────────
        lbl_estudiantes = QLabel("Selecciona un estudiante:")
        lbl_estudiantes.setFont(QFont("Arial", 11, QFont.Bold))
        layout.addWidget(lbl_estudiantes)
        
        self.tabla_estudiantes = QTableWidget()
        self.tabla_estudiantes.setColumnCount(5)
        self.tabla_estudiantes.setHorizontalHeaderLabels([
            "Código", "Nombre", "Email", "Carrera", "Ciclo"
        ])
        self.tabla_estudiantes.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla_estudiantes.setMaximumHeight(200)
        self.tabla_estudiantes.itemSelectionChanged.connect(self._on_estudiante_seleccionado)
        layout.addWidget(self.tabla_estudiantes)
        
        # ── REPORTE DETALLADO ──────────────────────────────────────
        lbl_reporte = QLabel("Detalle de Asistencia por Taller:")
        lbl_reporte.setFont(QFont("Arial", 11, QFont.Bold))
        layout.addWidget(lbl_reporte)
        
        self.tabla_reporte = QTableWidget()
        self.tabla_reporte.setColumnCount(8)
        self.tabla_reporte.setHorizontalHeaderLabels([
            "Taller", "Sesiones", "Presentes", "Justificados", 
            "Ausentes", "Asistencia %", "Umbral", "Estado"
        ])
        self.tabla_reporte.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        layout.addWidget(self.tabla_reporte, 1)
        
        # ── ESTADÍSTICAS ───────────────────────────────────────────
        stats_layout = QHBoxLayout()
        
        self.lbl_total_talleres = QLabel("Total Talleres: -")
        self.lbl_total_talleres.setFont(QFont("Arial", 10, QFont.Bold))
        stats_layout.addWidget(self.lbl_total_talleres)
        
        self.lbl_promedio = QLabel("Promedio Asistencia: -")
        self.lbl_promedio.setFont(QFont("Arial", 10, QFont.Bold))
        self.lbl_promedio.setStyleSheet("color: #534AB7;")
        stats_layout.addWidget(self.lbl_promedio)
        
        self.lbl_fecha_gen = QLabel("")
        self.lbl_fecha_gen.setStyleSheet("color: #999;")
        stats_layout.addWidget(self.lbl_fecha_gen)
        
        stats_layout.addStretch()
        layout.addLayout(stats_layout)
        
        # ── BOTONES ────────────────────────────────────────────────
        botones_layout = QHBoxLayout()
        
        self.btn_exportar_excel = QPushButton("📊 Exportar a Excel")
        self.btn_exportar_excel.clicked.connect(self._exportar_excel)
        self.btn_exportar_excel.setEnabled(False)
        botones_layout.addWidget(self.btn_exportar_excel)
        
        self.btn_exportar_pdf = QPushButton("📄 Exportar a PDF")
        self.btn_exportar_pdf.clicked.connect(self._exportar_pdf)
        self.btn_exportar_pdf.setEnabled(False)
        botones_layout.addWidget(self.btn_exportar_pdf)
        
        botones_layout.addStretch()
        
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.clicked.connect(self.reject)
        botones_layout.addWidget(btn_cerrar)
        
        layout.addLayout(botones_layout)
        
        self.setLayout(layout)
    
    def _crear_panel_filtros(self) -> QFrame:
        """Crea el panel de filtros."""
        panel = QFrame()
        panel.setStyleSheet("""
            QFrame { background-color: #f5f5f5; border-radius: 6px; padding: 10px; }
        """)
        
        layout = QVBoxLayout(panel)
        
        # Primera fila de filtros
        fila1 = QHBoxLayout()
        
        fila1.addWidget(QLabel("🔍 Buscar por nombre:"))
        self.inp_nombre = QLineEdit()
        self.inp_nombre.setPlaceholderText("Nombre o código...")
        self.inp_nombre.textChanged.connect(lambda: self._timer_busqueda.start())
        fila1.addWidget(self.inp_nombre)
        
        fila1.addWidget(QLabel("Ciclo:"))
        self.cmb_ciclo = QComboBox()
        self.cmb_ciclo.addItem("Todos", None)
        self.cmb_ciclo.currentIndexChanged.connect(self._timer_busqueda.start)
        fila1.addWidget(self.cmb_ciclo)
        
        layout.addLayout(fila1)
        
        # Segunda fila de filtros
        fila2 = QHBoxLayout()
        
        fila2.addWidget(QLabel("Carrera:"))
        self.cmb_carrera = QComboBox()
        self.cmb_carrera.addItem("Todas", None)
        self.cmb_carrera.currentIndexChanged.connect(self._timer_busqueda.start)
        fila2.addWidget(self.cmb_carrera)
        
        fila2.addWidget(QLabel("Taller:"))
        self.cmb_taller = QComboBox()
        self.cmb_taller.addItem("Todos", None)
        self.cmb_taller.currentIndexChanged.connect(self._timer_busqueda.start)
        fila2.addWidget(self.cmb_taller)
        
        btn_buscar = QPushButton("🔍 Buscar")
        btn_buscar.clicked.connect(self._ejecutar_busqueda)
        fila2.addWidget(btn_buscar)
        
        fila2.addStretch()
        layout.addLayout(fila2)
        
        return panel
    
    def _cargar_datos_iniciales(self) -> None:
        """Carga ciclos, carreras y talleres."""
        # Ciclos
        ciclos = self.reporte_service.listar_ciclos()
        for ciclo in ciclos:
            self.cmb_ciclo.addItem(ciclo["nombre"], ciclo["id"])
        
        # Carreras
        carreras = self.reporte_service.listar_carreras()
        for carrera in carreras:
            self.cmb_carrera.addItem(carrera["nombre"], carrera["id"])
        
        # Talleres
        talleres = self.reporte_service.listar_talleres_activos()
        for taller in talleres:
            self.cmb_taller.addItem(
                f"{taller['codigo']} - {taller['nombre']}",
                taller["id"]
            )
        
        # Búsqueda inicial
        self._ejecutar_busqueda()
    
    def _ejecutar_busqueda(self) -> None:
        """Ejecuta la búsqueda con filtros aplicados."""
        nombre = self.inp_nombre.text() if self.inp_nombre.text() else None
        ciclo_id = self.cmb_ciclo.currentData()
        carrera_id = self.cmb_carrera.currentData()
        taller_id = self.cmb_taller.currentData()
        
        resultado = self.reporte_service.listar_estudiantes_filtrados(
            ciclo_id=ciclo_id,
            nombre=nombre,
            carrera_id=carrera_id,
            taller_id=taller_id
        )
        
        if not resultado.ok:
            QMessageBox.warning(self, "Error", resultado.mensaje)
            return
        
        self.estudiantes_filtrados = resultado.lista or []
        self._mostrar_estudiantes()
    
    def _mostrar_estudiantes(self) -> None:
        """Muestra estudiantes en tabla."""
        self.tabla_estudiantes.setRowCount(len(self.estudiantes_filtrados))
        
        for fila, est in enumerate(self.estudiantes_filtrados):
            self.tabla_estudiantes.setItem(fila, 0, QTableWidgetItem(est["codigo"]))
            self.tabla_estudiantes.setItem(fila, 1, QTableWidgetItem(est["nombre"]))
            self.tabla_estudiantes.setItem(fila, 2, QTableWidgetItem(est["email"]))
            self.tabla_estudiantes.setItem(fila, 3, QTableWidgetItem(est["carrera"]))
            self.tabla_estudiantes.setItem(fila, 4, QTableWidgetItem(est["ciclo"]))
    
    def _on_estudiante_seleccionado(self) -> None:
        """Cuando se selecciona un estudiante."""
        filas = self.tabla_estudiantes.selectedIndexes()
        if not filas:
            return
        
        fila = filas[0].row()
        estudiante_id = self.estudiantes_filtrados[fila]["id"]
        
        # Generar reporte
        resultado = self.reporte_service.obtener_reporte_estudiante(estudiante_id)
        
        if not resultado.ok:
            QMessageBox.critical(self, "Error", resultado.mensaje)
            return
        
        self.reporte_actual = resultado.datos
        self._mostrar_reporte()
        
        self.btn_exportar_excel.setEnabled(True)
        self.btn_exportar_pdf.setEnabled(True)
    
    def _mostrar_reporte(self) -> None:
        """Muestra el reporte en la tabla."""
        if not self.reporte_actual:
            return
        
        datos = self.reporte_actual
        talleres = datos["talleres"]
        
        # Tabla de talleres
        self.tabla_reporte.setRowCount(len(talleres))
        
        for fila, taller in enumerate(talleres):
            self.tabla_reporte.setItem(fila, 0, QTableWidgetItem(taller["taller_nombre"]))
            self.tabla_reporte.setItem(fila, 1, QTableWidgetItem(str(taller["total_sesiones"])))
            self.tabla_reporte.setItem(fila, 2, QTableWidgetItem(str(taller["presentes"])))
            self.tabla_reporte.setItem(fila, 3, QTableWidgetItem(str(taller["justificados"])))
            self.tabla_reporte.setItem(fila, 4, QTableWidgetItem(str(taller["ausentes"])))
            
            item_pct = QTableWidgetItem(f"{taller['porcentaje_asistencia']:.1f}%")
            item_pct.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tabla_reporte.setItem(fila, 5, item_pct)
            
            # Mostrar umbral (si está disponible)
            umbral = taller.get("umbral_asistencia", "N/A")
            self.tabla_reporte.setItem(fila, 6, QTableWidgetItem(str(umbral)))
            
            # Estado con color
            item_estado = QTableWidgetItem(taller["estado"])
            if "✅" in taller["estado"]:
                item_estado.setBackground(QColor("#d1fae5"))
            elif "⚠️" in taller["estado"]:
                item_estado.setBackground(QColor("#fef3c7"))
            else:
                item_estado.setBackground(QColor("#fee2e2"))
            self.tabla_reporte.setItem(fila, 7, item_estado)
        
        # Estadísticas
        stats = datos["estadisticas_generales"]
        self.lbl_total_talleres.setText(f"Total Talleres: {stats['total_talleres']}")
        self.lbl_promedio.setText(f"Promedio Asistencia: {stats['promedio_asistencia']:.1f}%")
        self.lbl_fecha_gen.setText(f"Generado: {stats['fecha_generacion']}")
    
    def _exportar_excel(self) -> None:
        """Exporta reporte a Excel."""
        if not self.reporte_actual:
            return
        
        archivo, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar reporte como Excel",
            f"Reporte_{self.reporte_actual['estudiante']['codigo']}.xlsx",
            "Excel (*.xlsx)"
        )
        
        if not archivo:
            return
        
        # Preparar datos para exportación
        datos_export = {
            "estudiante_nombre": self.reporte_actual["estudiante"]["nombre"],
            "estudiante_codigo": self.reporte_actual["estudiante"]["codigo"],
            "fecha_generacion": self.reporte_actual["estadisticas_generales"]["fecha_generacion"]
        }
        
        # Usar ExportarExcel con estructura personalizada
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"
            
            # Encabezado
            ws['A1'] = "REPORTE DE ESTUDIANTE"
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['A3'] = "Estudiante:"
            ws['B3'] = self.reporte_actual["estudiante"]["nombre"]
            ws['A4'] = "Código:"
            ws['B4'] = self.reporte_actual["estudiante"]["codigo"]
            
            # Tabla
            ws['A6'] = "Taller"
            ws['B6'] = "Sesiones"
            ws['C6'] = "Presentes"
            ws['D6'] = "Justificados"
            ws['E6'] = "Ausentes"
            ws['F6'] = "Asistencia %"
            ws['G6'] = "Estado"
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws[f'{col}6'].font = Font(bold=True, color="FFFFFF")
                ws[f'{col}6'].fill = PatternFill(start_color="534AB7", end_color="534AB7", fill_type="solid")
            
            row = 7
            for taller in self.reporte_actual["talleres"]:
                ws[f'A{row}'] = taller["taller_nombre"]
                ws[f'B{row}'] = taller["total_sesiones"]
                ws[f'C{row}'] = taller["presentes"]
                ws[f'D{row}'] = taller["justificados"]
                ws[f'E{row}'] = taller["ausentes"]
                ws[f'F{row}'] = f"{taller['porcentaje_asistencia']:.1f}%"
                ws[f'G{row}'] = taller["estado"]
                row += 1
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['F'].width = 15
            
            wb.save(archivo)
            QMessageBox.information(self, "Éxito", f"Reporte exportado a:\n{archivo}")
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")
    
    def _exportar_pdf(self) -> None:
        """Exporta reporte a PDF."""
        if not self.reporte_actual:
            return
        
        archivo, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar reporte como PDF",
            f"Reporte_{self.reporte_actual['estudiante']['codigo']}.pdf",
            "PDF (*.pdf)"
        )
        
        if not archivo:
            return
        
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            
            doc = SimpleDocTemplate(archivo, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()
            
            # Título
            story.append(Paragraph("REPORTE DE ESTUDIANTE", styles['Title']))
            story.append(Spacer(1, 0.2 * inch))
            
            # Datos
            est = self.reporte_actual["estudiante"]
            datos = f"""
            <b>Estudiante:</b> {est['nombre']}<br/>
            <b>Código:</b> {est['codigo']}<br/>
            <b>Carrera:</b> {est['carrera']}<br/>
            """
            story.append(Paragraph(datos, styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
            
            # Tabla
            tabla_datos = [["Taller", "Sesiones", "Presentes", "Asistencia %", "Estado"]]
            for t in self.reporte_actual["talleres"]:
                tabla_datos.append([
                    t["taller_nombre"],
                    str(t["total_sesiones"]),
                    str(t["presentes"]),
                    f"{t['porcentaje_asistencia']:.1f}%",
                    t["estado"]
                ])
            
            tabla = Table(tabla_datos)
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#534AB7')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            
            story.append(tabla)
            doc.build(story)
            
            QMessageBox.information(self, "Éxito", f"PDF exportado a:\n{archivo}")
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")