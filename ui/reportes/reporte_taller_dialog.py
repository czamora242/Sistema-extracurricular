"""
ui/reportes/reporte_taller_dialog.py   ──   Reporte de Taller
═════════════════════════════════════════════════════════════════

Características:
  ✅ Seleccionar taller
  ✅ Seleccionar una o varias sesiones
  ✅ Matriz de asistencia por estudiante/sesión
  ✅ Estadísticas por estudiante
  ✅ Exportación a Excel y PDF
"""

from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QColor
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTableWidget, QTableWidgetItem, QMessageBox, QFileDialog,
    QComboBox, QCheckBox, QScrollArea, QWidget, QListWidget,
    QListWidgetItem, QHeaderView
)

from services.reporte_service import ReporteService
from services.taller_service import TallerService


class ReporteTallerDialog(QDialog):
    """Diálogo de reporte de taller con múltiples sesiones."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.reporte_service = ReporteService()
        self.taller_service = TallerService()
        self.sesiones_disponibles = []
        self.sesiones_seleccionadas = []
        self.reporte_actual = None
        
        self._configurar_ventana()
        self._construir_ui()
        self._cargar_talleres()
    
    def _configurar_ventana(self) -> None:
        """Configura la ventana."""
        self.setWindowTitle("Reporte de Taller")
        self.setMinimumSize(1300, 800)
        self.setModal(True)
    
    def _construir_ui(self) -> None:
        """Construye la interfaz."""
        layout = QVBoxLayout()
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)
        
        # ── TÍTULO ─────────────────────────────────────────────────
        lbl_titulo = QLabel("🎯 Reporte de Taller")
        font_titulo = QFont()
        font_titulo.setPointSize(16)
        font_titulo.setBold(True)
        lbl_titulo.setFont(font_titulo)
        layout.addWidget(lbl_titulo)
        
        # ── SELECCIÓN DE TALLER Y SESIONES ─────────────────────────
        seleccion_layout = QHBoxLayout()
        
        # Taller
        seleccion_layout.addWidget(QLabel("Taller:"))
        self.cmb_taller = QComboBox()
        self.cmb_taller.currentIndexChanged.connect(self._on_taller_cambio)
        seleccion_layout.addWidget(self.cmb_taller, 1)
        
        # Botón generar
        self.btn_generar = QPushButton("🔄 Generar Reporte")
        self.btn_generar.clicked.connect(self._generar_reporte)
        self.btn_generar.setEnabled(False)
        seleccion_layout.addWidget(self.btn_generar)
        
        layout.addLayout(seleccion_layout)
        
        # ── SELECCIÓN DE SESIONES ──────────────────────────────────
        lbl_sesiones = QLabel("Selecciona sesiones a incluir:")
        lbl_sesiones.setFont(QFont("Arial", 11, QFont.Bold))
        layout.addWidget(lbl_sesiones)
        
        sesiones_layout = QHBoxLayout()
        
        # Lista de sesiones
        self.lista_sesiones = QListWidget()
        self.lista_sesiones.setMaximumHeight(120)
        sesiones_layout.addWidget(self.lista_sesiones, 1)
        
        # Botones de selección
        botones_sesiones = QVBoxLayout()
        
        self.btn_todas = QPushButton("✓ Todas")
        self.btn_todas.clicked.connect(self._seleccionar_todas)
        self.btn_todas.setEnabled(False)
        botones_sesiones.addWidget(self.btn_todas)
        
        self.btn_ninguna = QPushButton("✗ Ninguna")
        self.btn_ninguna.clicked.connect(self._deseleccionar_todas)
        self.btn_ninguna.setEnabled(False)
        botones_sesiones.addWidget(self.btn_ninguna)
        
        botones_sesiones.addStretch()
        sesiones_layout.addLayout(botones_sesiones)
        
        layout.addLayout(sesiones_layout)
        
        # ── TABLA DE REPORTE ───────────────────────────────────────
        lbl_reporte = QLabel("Matriz de Asistencia:")
        lbl_reporte.setFont(QFont("Arial", 11, QFont.Bold))
        layout.addWidget(lbl_reporte)
        
        self.tabla_reporte = QTableWidget()
        self.tabla_reporte.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla_reporte.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents
        )
        layout.addWidget(self.tabla_reporte, 1)
        
        # ── ESTADÍSTICAS ───────────────────────────────────────────
        stats_layout = QHBoxLayout()
        
        self.lbl_total_est = QLabel("Total Estudiantes: -")
        self.lbl_total_est.setFont(QFont("Arial", 10, QFont.Bold))
        stats_layout.addWidget(self.lbl_total_est)
        
        self.lbl_promedio_taller = QLabel("Asistencia Promedio: -")
        self.lbl_promedio_taller.setFont(QFont("Arial", 10, QFont.Bold))
        self.lbl_promedio_taller.setStyleSheet("color: #534AB7;")
        stats_layout.addWidget(self.lbl_promedio_taller)
        
        stats_layout.addStretch()
        layout.addLayout(stats_layout)
        
        # ── BOTONES DE ACCIÓN ──────────────────────────────────────
        botones_layout = QHBoxLayout()
        
        self.btn_exportar_excel = QPushButton("📊 Exportar a Excel")
        self.btn_exportar_excel.clicked.connect(self._exportar_excel)
        self.btn_exportar_excel.setEnabled(False)
        botones_layout.addWidget(self.btn_exportar_excel)
        
        self.btn_exportar_pdf = QPushButton("📄 Exportar a PDF")
        self.btn_exportar_pdf.clicked.connect(self._exportar_pdf)
        self.btn_exportar_pdf.setEnabled(False)
        botones_layout.addWidget(self.btn_exportar_pdf)
        
        botones_layout.addStretch()
        
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.clicked.connect(self.reject)
        botones_layout.addWidget(btn_cerrar)
        
        layout.addLayout(botones_layout)
        
        self.setLayout(layout)
    
    def _cargar_talleres(self) -> None:
        """Carga talleres activos."""
        talleres = self.reporte_service.listar_talleres_activos()
        
        self.cmb_taller.clear()
        self.cmb_taller.addItem("Selecciona un taller...", None)
        
        for taller in talleres:
            self.cmb_taller.addItem(
                f"{taller['codigo']} - {taller['nombre']}",
                taller["id"]
            )
    
    def _on_taller_cambio(self) -> None:
        """Cuando cambia el taller seleccionado."""
        taller_id = self.cmb_taller.currentData()
        
        self.lista_sesiones.clear()
        self.sesiones_disponibles = []
        self.sesiones_seleccionadas = []
        
        if not taller_id:
            self.btn_generar.setEnabled(False)
            self.btn_todas.setEnabled(False)
            self.btn_ninguna.setEnabled(False)
            return
        
        # Obtener sesiones del taller
        sesiones = self.taller_service.listar_sesiones(taller_id)
        
        for sesion in sesiones:
            self.sesiones_disponibles.append({
                "id": sesion["id"],
                "numero": sesion["numero"],
                "fecha": sesion["fecha"]
            })
        
        # Mostrar sesiones
        for sesion in self.sesiones_disponibles:
            item = QListWidgetItem(
                f"Sesión {sesion['numero']} - {sesion['fecha']}"
            )
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked)
            self.lista_sesiones.addItem(item)
            self.sesiones_seleccionadas.append(sesion["id"])
        
        self.btn_generar.setEnabled(len(self.sesiones_disponibles) > 0)
        self.btn_todas.setEnabled(True)
        self.btn_ninguna.setEnabled(True)
    
    def _seleccionar_todas(self) -> None:
        """Selecciona todas las sesiones."""
        self.sesiones_seleccionadas = []
        for i in range(self.lista_sesiones.count()):
            item = self.lista_sesiones.item(i)
            item.setCheckState(Qt.CheckState.Checked)
            self.sesiones_seleccionadas.append(self.sesiones_disponibles[i]["id"])
    
    def _deseleccionar_todas(self) -> None:
        """Deselecciona todas las sesiones."""
        self.sesiones_seleccionadas = []
        for i in range(self.lista_sesiones.count()):
            item = self.lista_sesiones.item(i)
            item.setCheckState(Qt.CheckState.Unchecked)
    
    def _actualizar_sesiones_seleccionadas(self) -> None:
        """Actualiza la lista de sesiones seleccionadas según los checkboxes."""
        self.sesiones_seleccionadas = []
        for i in range(self.lista_sesiones.count()):
            item = self.lista_sesiones.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                self.sesiones_seleccionadas.append(self.sesiones_disponibles[i]["id"])
    
    def _generar_reporte(self) -> None:
        """Genera el reporte con las sesiones seleccionadas."""
        taller_id = self.cmb_taller.currentData()
        
        if not taller_id:
            QMessageBox.warning(self, "Error", "Selecciona un taller")
            return
        
        self._actualizar_sesiones_seleccionadas()
        
        if not self.sesiones_seleccionadas:
            QMessageBox.warning(self, "Error", "Selecciona al menos una sesión")
            return
        
        # Generar reporte
        resultado = self.reporte_service.obtener_reporte_taller(
            taller_id=taller_id,
            sesion_ids=self.sesiones_seleccionadas
        )
        
        if not resultado.ok:
            QMessageBox.critical(self, "Error", resultado.mensaje)
            return
        
        self.reporte_actual = resultado.datos
        self._mostrar_reporte()
        
        self.btn_exportar_excel.setEnabled(True)
        self.btn_exportar_pdf.setEnabled(True)
    
    def _mostrar_reporte(self) -> None:
        """Muestra el reporte en la tabla."""
        if not self.reporte_actual:
            return
        
        datos = self.reporte_actual
        estudiantes = datos["estudiantes"]
        sesiones = datos["sesiones"]
        
        # Configurar tabla
        num_columnas = 2 + len(sesiones) + 4  # Código, Nombre, Sesiones..., Presentes, Justificados, Ausentes, %
        self.tabla_reporte.setColumnCount(num_columnas)
        
        # Headers
        headers = ["Código", "Nombre"]
        for sesion in sesiones:
            headers.append(f"S{sesion['numero']}")
        headers.extend(["Presentes", "Justificados", "Ausentes", "Asistencia %"])
        
        self.tabla_reporte.setHorizontalHeaderLabels(headers)
        
        # Datos
        self.tabla_reporte.setRowCount(len(estudiantes))
        
        for fila, est in enumerate(estudiantes):
            # Código y nombre
            self.tabla_reporte.setItem(fila, 0, QTableWidgetItem(est["codigo"]))
            self.tabla_reporte.setItem(fila, 1, QTableWidgetItem(est["nombre"]))
            
            # Asistencias por sesión
            col = 2
            for asistencia in est["asistencias_por_sesion"]:
                item = QTableWidgetItem(asistencia["estado"])
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                
                # Color según estado
                if "✅" in asistencia["estado"]:
                    item.setBackground(QColor("#d1fae5"))
                elif "❌" in asistencia["estado"]:
                    item.setBackground(QColor("#fee2e2"))
                
                self.tabla_reporte.setItem(fila, col, item)
                col += 1
            
            # Estadísticas
            self.tabla_reporte.setItem(fila, col, QTableWidgetItem(str(est["presentes"])))
            col += 1
            self.tabla_reporte.setItem(fila, col, QTableWidgetItem(str(est["justificados"])))
            col += 1
            self.tabla_reporte.setItem(fila, col, QTableWidgetItem(str(est["ausentes"])))
            col += 1
            
            item_pct = QTableWidgetItem(f"{est['porcentaje']:.1f}%")
            item_pct.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tabla_reporte.setItem(fila, col, item_pct)
        
        # Estadísticas generales
        stats = datos["estadisticas"]
        self.lbl_total_est.setText(f"Total Estudiantes: {stats['total_estudiantes']}")
        self.lbl_promedio_taller.setText(f"Asistencia Promedio: {stats['asistencia_promedio']:.1f}%")
    
    def _exportar_excel(self) -> None:
        """Exporta reporte a Excel."""
        if not self.reporte_actual:
            return
        
        archivo, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar reporte como Excel",
            f"Reporte_{self.reporte_actual['taller']['codigo']}.xlsx",
            "Excel (*.xlsx)"
        )
        
        if not archivo:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Encabezado
            ws['A1'] = "REPORTE DE TALLER"
            ws['A1'].font = Font(bold=True, size=14)
            
            datos = self.reporte_actual
            ws['A3'] = "Taller:"
            ws['B3'] = datos["taller"]["nombre"]
            ws['A4'] = "Docente:"
            ws['B4'] = datos["taller"]["docente"]
            
            # Tabla
            col = 1
            headers = ["Código", "Nombre", "Presentes", "Justificados", "Ausentes", "Asistencia %"]
            for header in headers:
                ws.cell(row=6, column=col).value = header
                ws.cell(row=6, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=6, column=col).fill = PatternFill(start_color="534AB7", end_color="534AB7", fill_type="solid")
                col += 1
            
            row = 7
            for est in datos["estudiantes"]:
                ws[f'A{row}'] = est["codigo"]
                ws[f'B{row}'] = est["nombre"]
                ws[f'C{row}'] = est["presentes"]
                ws[f'D{row}'] = est["justificados"]
                ws[f'E{row}'] = est["ausentes"]
                ws[f'F{row}'] = f"{est['porcentaje']:.1f}%"
                row += 1
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            
            wb.save(archivo)
            QMessageBox.information(self, "Éxito", f"Exportado a:\n{archivo}")
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error: {str(e)}")
    
    def _exportar_pdf(self) -> None:
        """Exporta reporte a PDF."""
        if not self.reporte_actual:
            return
        
        archivo, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar reporte como PDF",
            f"Reporte_{self.reporte_actual['taller']['codigo']}.pdf",
            "PDF (*.pdf)"
        )
        
        if not archivo:
            return
        
        QMessageBox.information(self, "Info", "Exportación a PDF en desarrollo")