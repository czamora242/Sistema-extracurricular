"""
utils/exportar_excel.py   ──   Exportación a Excel
═════════════════════════════════════════════════

Utilidades para exportar datos a archivos Excel (.xlsx).
Requiere: pip install openpyxl

USO:
    from utils.exportar_excel import ExportarExcel
    
    ExportarExcel.lista_aptos(
        datos=resultado.datos,
        estudiantes=resultado.lista,
        ruta="/ruta/al/archivo.xlsx"
    )
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportarExcel:
    """Exportación a Excel para varios módulos."""

    @staticmethod
    def lista_aptos(datos: Dict, estudiantes: Dict[str, List], 
                    ruta: str = None) -> bool:
        """
        Exporta lista de aptos a Excel con formato profesional.
        
        PARÁMETROS:
            datos: dict con información del taller
            estudiantes: dict con "aptos" y "desaptos"
            ruta: ruta de descarga (si None, usa Desktop)
        
        RETORNA:
            bool: True si éxito
        """
        if not OPENPYXL_AVAILABLE:
            print("❌ openpyxl no está instalado. Usa: pip install openpyxl")
            return False

        try:
            # ── Crear libro ──────────────────────────────────────────
            wb = Workbook()
            ws = wb.active
            ws.title = "Lista de Aptos"

            # ── Encabezado ───────────────────────────────────────────
            ws['A1'] = "LISTA DE APTOS"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = Alignment(horizontal='center')

            # Información del taller
            row = 3
            ws[f'A{row}'] = "Taller:"
            ws[f'B{row}'] = datos.get('taller_nombre', 'N/A')
            ws[f'B{row}'].font = Font(bold=True)

            row += 1
            ws[f'A{row}'] = "Código:"
            ws[f'B{row}'] = datos.get('taller_codigo', 'N/A')

            row += 1
            ws[f'A{row}'] = "Docente:"
            ws[f'B{row}'] = datos.get('docente', 'N/A')

            row += 1
            ws[f'A{row}'] = "Ciclo:"
            ws[f'B{row}'] = datos.get('ciclo', 'N/A')

            row += 1
            ws[f'A{row}'] = "Umbral:"
            ws[f'B{row}'] = f"{datos.get('umbral', 0)}%"

            row += 1
            ws[f'A{row}'] = "Fecha de generación:"
            ws[f'B{row}'] = datos.get('fecha_generacion', datetime.now().strftime("%d/%m/%Y"))

            # ── Tabla de datos ───────────────────────────────────────
            row = 10
            
            # Encabezados de tabla
            headers = ["Nombre", "DNI", "Código", "Carrera", "Asistencia", "Umbral", "Estado"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="534AB7", end_color="534AB7", 
                                       fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # ── Datos de APTOS (en verde) ────────────────────────────
            row += 1
            aptos = estudiantes.get('aptos', [])
            
            for estudiante in aptos:
                ws.cell(row=row, column=1).value = estudiante.get('nombre_completo')
                ws.cell(row=row, column=2).value = estudiante.get('dni', 'N/A')
                ws.cell(row=row, column=3).value = estudiante.get('codigo_estudiantil', 'N/A')
                ws.cell(row=row, column=4).value = estudiante.get('carrera', 'N/A')
                
                asistencia_str = f"{estudiante.get('asistencia_porcentaje', 0):.1f}%"
                ws.cell(row=row, column=5).value = asistencia_str
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
                
                ws.cell(row=row, column=6).value = f"{datos.get('umbral', 0)}%"
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
                
                ws.cell(row=row, column=7).value = "APTO"
                ws.cell(row=row, column=7).font = Font(bold=True, color="1D9E75")
                ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
                
                # Fondo verde claro
                green_fill = PatternFill(start_color="E8F8F2", end_color="E8F8F2",
                                        fill_type="solid")
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = green_fill
                
                row += 1

            # ── Datos de DESAPTOS (en rojo) ──────────────────────────
            desaptos = estudiantes.get('desaptos', [])
            
            for estudiante in desaptos:
                ws.cell(row=row, column=1).value = estudiante.get('nombre_completo')
                ws.cell(row=row, column=2).value = estudiante.get('dni', 'N/A')
                ws.cell(row=row, column=3).value = estudiante.get('codigo_estudiantil', 'N/A')
                ws.cell(row=row, column=4).value = estudiante.get('carrera', 'N/A')
                
                asistencia_str = f"{estudiante.get('asistencia_porcentaje', 0):.1f}%"
                ws.cell(row=row, column=5).value = asistencia_str
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
                
                ws.cell(row=row, column=6).value = f"{datos.get('umbral', 0)}%"
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
                
                ws.cell(row=row, column=7).value = "DESAPTO"
                ws.cell(row=row, column=7).font = Font(bold=True, color="C0392B")
                ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
                
                # Fondo rojo claro
                red_fill = PatternFill(start_color="FDF0EE", end_color="FDF0EE",
                                      fill_type="solid")
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = red_fill
                
                row += 1

            # ── Ajustar anchos de columna ────────────────────────────
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12

            # ── Resumen al final ─────────────────────────────────────
            row += 2
            ws.cell(row=row, column=1).value = "RESUMEN"
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            row += 1
            ws.cell(row=row, column=1).value = f"Aptos: {len(aptos)}"
            ws.cell(row=row, column=2).value = f"Desaptos: {len(desaptos)}"
            ws.cell(row=row, column=3).value = f"Total: {len(aptos) + len(desaptos)}"

            # ── Determinar ruta de guardado ──────────────────────────
            if not ruta:
                # Guardar en Desktop o Documents
                home = Path.home()
                desktop = home / "Desktop"
                if not desktop.exists():
                    desktop = home / "Documents"
                if not desktop.exists():
                    desktop = home
                
                fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
                nombre_taller = datos.get('taller_nombre', 'taller').replace(' ', '_')
                ruta = desktop / f"Lista_Aptos_{nombre_taller}_{fecha}.xlsx"
            else:
                ruta = Path(ruta)

            # ── Guardar archivo ─────────────────────────────────────
            ruta.parent.mkdir(parents=True, exist_ok=True)
            wb.save(ruta)

            print(f"✅ Archivo guardado: {ruta}")
            return True

        except Exception as e:
            print(f"❌ Error al exportar: {e}")
            return False