"""
utils/docente_excel_importer.py   ──   Importar Docentes desde Excel
════════════════════════════════════════════════════════════════════════

¿QUÉ HACE?
  • Lee archivo Excel con datos de docentes
  • Valida cada fila
  • Detecta duplicados (DNI)
  • Inserta docentes en la BD
  • Retorna reporte con resultados

FORMATO EXCEL ESPERADO:
  Columna A: DNI
  Columna B: Nombres
  Columna C: Apellidos
  Columna D: Especialidad (opcional)
  Columna E: Email Institucional (opcional)
  Columna F: Teléfono (opcional)

EJEMPLO:
  DNI       | Nombres | Apellidos | Especialidad | Email         | Teléfono
  12345678  | Juan    | García    | Sistemas     | juan@...      | 999000000
  87654321  | Maria   | López     | Contabilidad | maria@...     | 999111111
"""

from typing import Dict, List, Any
from openpyxl import load_workbook
from services.docente_service import DocenteService


class DocenteExcelImporter:
    """Importa docentes desde archivo Excel."""

    def importar(
        self,
        ruta_archivo: str,
        usuario_creador_id: int = None,
        hoja: int = 0
    ) -> Dict[str, Any]:
        """
        Importa docentes desde Excel.

        PARÁMETROS:
          ruta_archivo: ruta al archivo .xlsx
          usuario_creador_id: ID del usuario que está importando
          hoja: índice de la hoja a leer (default 0 = primera)

        RETORNA:
          Dict con:
            - ok: bool
            - mensaje: str
            - exitosos: int
            - errores: int
            - duplicados: int
            - detalles: List[str]
        """

        resultado = {
            "ok": False,
            "mensaje": "",
            "exitosos": 0,
            "errores": 0,
            "duplicados": 0,
            "detalles": []
        }

        try:
            # Cargar workbook
            wb = load_workbook(ruta_archivo)
            ws = wb.worksheets[hoja]

            # Validar que tenga datos
            if ws.max_row < 2:
                resultado["mensaje"] = "El archivo no contiene datos."
                return resultado

            # Procesar filas
            dni_procesados = set()

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extraer datos
                    dni = str(row[0] or "").strip()
                    nombres = str(row[1] or "").strip()
                    apellidos = str(row[2] or "").strip()
                    especialidad = str(row[3] or "").strip() if row[3] else None
                    email = str(row[4] or "").strip().lower() if row[4] else None
                    telefono = str(row[5] or "").strip() if row[5] else None

                    # Validaciones básicas
                    if not dni or not nombres or not apellidos:
                        resultado["detalles"].append(
                            f"Fila {row_idx}: Faltan datos obligatorios (DNI, Nombres, Apellidos)"
                        )
                        resultado["errores"] += 1
                        continue

                    # Validar DNI único en el archivo
                    if dni in dni_procesados:
                        resultado["detalles"].append(
                            f"Fila {row_idx}: DNI '{dni}' duplicado en el archivo"
                        )
                        resultado["duplicados"] += 1
                        continue

                    dni_procesados.add(dni)

                    # Validar DNI único en BD
                    res_existe = DocenteService.obtener_por_dni(dni)
                    if res_existe.ok:
                        resultado["detalles"].append(
                            f"Fila {row_idx}: DNI '{dni}' ya existe en la BD"
                        )
                        resultado["duplicados"] += 1
                        continue

                    # Crear docente
                    res_crear = DocenteService.crear(
                        dni=dni,
                        nombres=nombres,
                        apellidos=apellidos,
                        especialidad=especialidad,
                        email_institucional=email,
                        telefono=telefono,
                        usuario_id=None,  # Sin usuario
                        usuario_creador_id=usuario_creador_id
                    )

                    if res_crear.ok:
                        resultado["exitosos"] += 1
                    else:
                        resultado["detalles"].append(
                            f"Fila {row_idx}: {res_crear.mensaje}"
                        )
                        resultado["errores"] += 1

                except Exception as e:
                    resultado["detalles"].append(
                        f"Fila {row_idx}: Error al procesar - {str(e)}"
                    )
                    resultado["errores"] += 1

            # Determinar éxito
            resultado["ok"] = resultado["exitosos"] > 0
            resultado["mensaje"] = (
                f"Importación completada: "
                f"{resultado['exitosos']} exitosos, "
                f"{resultado['errores']} errores, "
                f"{resultado['duplicados']} duplicados"
            )

            return resultado

        except FileNotFoundError:
            resultado["mensaje"] = "El archivo no existe."
            return resultado

        except Exception as e:
            resultado["mensaje"] = f"Error al leer el archivo: {str(e)}"
            return resultado

    @staticmethod
    def generar_plantilla(ruta_salida: str = "plantilla_docentes.xlsx") -> bool:
        """
        Genera una plantilla Excel vacía para importar docentes.

        PARÁMETROS:
          ruta_salida: donde guardar la plantilla

        RETORNA:
          True si se creó exitosamente, False si error
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Docentes"

            # Encabezados
            headers = [
                "DNI",
                "Nombres",
                "Apellidos",
                "Especialidad",
                "Email Institucional",
                "Teléfono"
            ]

            # Escribir encabezados con formato
            header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Agregar fila de ejemplo
            ejemplo = ["12345678", "Juan", "García", "Sistemas", "juan@unab.edu.pe", "+51999000000"]
            for col, valor in enumerate(ejemplo, start=1):
                cell = ws.cell(row=2, column=col)
                cell.value = valor
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Ajustar ancho de columnas
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 25
            ws.column_dimensions["F"].width = 15

            # Guardar
            wb.save(ruta_salida)
            return True

        except Exception as e:
            print(f"Error al generar plantilla: {str(e)}")
            return False
